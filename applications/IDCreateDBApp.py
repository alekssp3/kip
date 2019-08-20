'''Генератор базы из журналов входного контроля
На на выходе получается xlsx файл формата
project-emid obj cert
Сначала необходимо сгенерировать комплекты используя
IDFolderInitApp'''

from pathlib import Path
from kip.utils.Config import Config
from kip.core.utils import make_absolute
import re
import openpyxl
import datetime
from testing import ICJData


app_path = Path(__file__).parent


def simple_normilize(lst:list):
    out = []
    for l in lst:
        out.append(' '.join(l.split()))
    return out


def getDataDict(data):
    pattern = r'\d{5} *[eEеЕ]'
    regex = re.compile(pattern)
    out = {}
    for d in data:
        if d is not None:
            proj = str(d.book).split('\\')[-3].split('_')[0]
            for _d in d.data:
                for match in regex.finditer(_d.obj):
                    s = match.start()
                    e = match.end()
                    emid = _d.obj[s:e-1]
                    obj = _d.obj[0:s].strip()
                    cert = _d.cert
                    proj_emid = proj + '-' + emid + 'E'
                    # print(f'{proj_emid} -> {obj} -> {cert}')
                    if proj_emid in out.keys():
                        out[proj_emid][0].append(obj)
                        out[proj_emid][1].append(cert)
                    else:
                        out[proj_emid] = [[obj], [cert]]
    for d in out:
        out[d] = [list(set(simple_normilize(out[d][0]))), list(set(simple_normilize(out[d][1])))]
    return out


def createDB(name, find_path=None, save_path=None):
    handle = 'proj-emid obj cert'
    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = name
    for i, c in enumerate(handle.split()):
        ws.cell(1, i+1).value = c
    _dict = getDataDict(ICJData.get(find_path))
    for row, d in enumerate(_dict):
        ws.cell(row+2, 1).value = d
        obj = '\n'.join(_dict[d][0]).strip()
        ws.cell(row+2, 2).value = obj
        cert = '\n'.join(_dict[d][1]).strip()
        ws.cell(row+2, 3).value = cert
    ws.auto_filter.ref = 'A1:C'+str(row+1)
    wb.save(save_path.joinpath(name + '.xlsx'))
    wb.close()


def main(*args, **kwargs):
    config = Config(path=make_absolute(app_path, 'IDCreateDB.conf'))
    # print(make_absolute(app_path, 'IDCreateDB.conf'))
    path_to_save_db = make_absolute(app_path, config.PATH_TO_SAVE_DB)
    path_to_find_comp = make_absolute(app_path, config.PATH_TO_FIND_COMP)
    if config.DEFAULT_NAME is None or config.DEFAULT_NAME == '':
        defaultname = 'ICJ_DB_{:%d%m%y}'.format(datetime.datetime.today())
    else:
        defaultname = config.DEFAULT_NAME
    # print(defaultname)
    createDB(defaultname, path_to_find_comp, path_to_save_db)


if __name__ == '__main__':
    main()