import re
import openpyxl
import xlwings as xw
import xlrd

from ...Scanner.Scanner import Scanner
from ..Splitter import Splitter
from ...utils.data import load, save
from .excel_utils import get_data_from_icj, get_workbook, get_needed_sheet
from ..Structures import PROJECT_STRUCTURES, get_structure
from ..Utils import pne
# from collections import namedtuple

class WrongFileFormat(Exception):
    pass


def convert_to_excel(filename):
    """Data format:
    title
    description
    data
    data

    title2
    ...

    title3
    ..."""
    buffer = load(filename, 'raw')
    out = []
    spl = Splitter()
    if re.search(r'\n{2}', buffer) is None:
        raise WrongFileFormat
    sub_buf = spl.by_dbl_nl(buffer)
    for line in sub_buf:
        buf = spl.by_nl(line)
        for i in buf[2:]:
            out.append('\t'.join((buf[0], i, buf[1])))
    if len(out) != 0:
        # TODO: add history for aborting this operation
        save(out, filename)
    else:
        print('Buffer is empty')


# def convert_from_excel(filename):
#     """Data format:
#     row1 row2 row3 row4 ...
#     row1 row2 row3 row4 ...
#     row1 row2 row3 row4 ...
#     ..."""
#     buffer = load(filename, 'raw')
#     out = []
#     if re.search(r'\n{2}', buffer) == None:
#         raise Exception('Wrong file format')
#     sub_buf = re.split(r'\n\n', buffer)
#     for line in sub_buf:
#         buf = re.split(r'\n', line)
#         for i in buf[2:]:
#             out.append('\t'.join((buf[0], i, buf[1])))
#     if len(out) != 0:
#         save(out, filename)
#     else:
#         print('Buffer is empty')


def timesheetsum(filename):
    "Подсчет человекочасов в книге по форме RHI"
    wb = openpyxl.load_workbook(filename)
    ws = wb.sheetnames
    out = []
    for sh in ws[1:]:
        _sh = wb.get_sheet_by_name(sh)
        s = sum([_sh['h'+ str(j)].value for j in range(9, 48) if _sh['h'+ str(j)].value is not None])
        out.append(f'{sh}\t{s}')
    return out

# def get_selected_data():
#     return xw.books.active.selection.value

# def _swap_data(data, x, y):
#     data[x], data[y] = data[y], data[x]
#     return data

# def swap_selected_columns(selection_indexes):
#     data = get_selected_data()
#     i, j = selection_indexes
#     out = [_swap_data(d, i, j) for d in data]
#     return out


class ICJ:
    'Incoming control journal'
    def __init_(self):
        self.buf

    def copy(self):
        self.buf = xw.books.active.selection.value

    def paste(self, buf=None):
        if buf is None:
            xw.books.active.selection.value = self.buf
        else:
            xw.books.active.selection.value = buf

    def check(self):
        out = []
        for id, item in enumerate(self.buf):
            if len(item[0]) > 0:
                skl = item[0].split()
            else:
                skl = ['', '']
            pr = item[1].split()
            if len(pr) < 2:
                pr.append('')
            if skl[1] != pr[1]:
                print(f'{item}: {skl[1]} != {pr[1]}')
            out.append([' '.join((skl[0], pr[1])), ' '.join(pr)])
            if pr[1] == 'м':
                xw.books.active.sheets.active.range('n'+str(id+4)).value = ''
        return out

    def smart_check(self):
        # check_for_quality
        # check_for_unit
        pass


def check_and_close():
    icj = ICJ()
    icj.copy()
    icj.paste(icj.check())
    xw.Book('Кузня ЖВК KIPEM.xlsb').macro('AutoFilter')()


def zbs(start, end=None):
    if end is None:
        end = start
    wb = xw.Book('Кузня ЖВК KIPEM.xlsb')
    ws = wb.sheets[1]
    wr = ws.range('t1')
    for i in range(start, end+1):
        wr.value = i
        wb.macro('MakeJournal')()
        print(xw.books.active.name)
        check_and_close()


def numerator2(arr):
    'String numerator by page count'
    out = []
    page_sum = 0
    for data in arr:
        page_count = int(data)
        if page_count == 1:
            out.append(str(page_count + page_sum))
        else:
            out.append(str(page_sum + 1) + ' - ' + str(page_count + page_sum))
        page_sum = page_count + page_sum
    return out


def numerator3(arr):
    'String numerator by page count and index'
    out = []
    page_sum = 0
    prev_data = ''
    for data in arr:
        page_count = abs(int(data[0]))
        cur_data = data[1]
        if cur_data != prev_data:
            page_sum = 0
        if page_count == 0:
            page_count = 1
        if page_count == 1:
            out.append(str(page_sum + 1))
        else:
            out.append(str(page_sum + 1) + ' - ' + str(page_count + page_sum))
        page_sum = page_count + page_sum
        prev_data = cur_data
    return out


icj = ICJ()


def autonum_on_place():
    'Autonumeration second column'
    icj.copy()
    o = numerator3(icj.buf)
    o = [[i] for i in o]
    sel = xw.books.active.selection
    sel.offset(0, 1).select()
    # sel.offset(0, -1).select()
    icj.paste(o)


def check_data(data):
    errors = []
    for d in data:
        if d == []:
            errors.append(d.name)
    pne(errors, 'No data')
