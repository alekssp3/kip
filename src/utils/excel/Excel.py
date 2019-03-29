import re
import openpyxl

from ..Splitter import Splitter
from ..Utils import load, save
from ..dxf.DXF import WrongFileFormat


def convert_to_excel(filename):
    """Data format:
    title
    description
    data
    data

    title2
    ...

    title3
    ..."""
    buffer = load(filename, 'raw')
    out = []
    spl = Splitter()
    if re.search(r'\n{2}', buffer) == None:
        raise WrongFileFormat
    sub_buf = spl.by_dbl_nl(buffer)
    for line in sub_buf:
        buf = spl.by_nl(line)
        for i in buf[2:]:
            out.append('\t'.join((buf[0], i, buf[1])))
    if len(out) != 0:
        # TODO: add history for aborting this operation
        save(out, filename)
    else:
        print('Buffer is empty')


# def convert_from_excel(filename):
#     """Data format:
#     row1 row2 row3 row4 ...
#     row1 row2 row3 row4 ...
#     row1 row2 row3 row4 ...
#     ..."""
#     buffer = load(filename, 'raw')
#     out = []
#     if re.search(r'\n{2}', buffer) == None:
#         raise Exception('Wrong file format')
#     sub_buf = re.split(r'\n\n', buffer)
#     for line in sub_buf:
#         buf = re.split(r'\n', line)
#         for i in buf[2:]:
#             out.append('\t'.join((buf[0], i, buf[1])))
#     if len(out) != 0:
#         save(out, filename)
#     else:
#         print('Buffer is empty')


def timesheetsum(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.sheetnames
    out = []
    for sh in ws[1:]:
        s = sum([sh['b'+ str(j)].value for j in range(4, 10) if sh['b'+ str(j)].value is not None])
        out.append(f'{sh}\t{s}')
    return out