from kip.utils.Structures import PROJECT_STRUCTURES, get_structure
from kip.utils.excel.excel_utils import get_needed_sheet, get_workbook, get_data_from_icj
import xlwings
from kip.Scanner.Scanner import Scanner
import re
import openpyxl


def get(path):
    scanner = Scanner().scan(path)
    return getICJDatas(scanner.get_struct())


def getICJData(path):
    # data = []
    # scanner = Scanner()
    # scanner.scan(path)
    # st = scanner.get_struct()
    # files = list(filter(lambda a: a.ext=='.xlsx' and 'ЖВК' in a.name and '~' not in a.name, sorted(st, key=lambda a: a.date.c)))
    # workbooks = get_workbooks([i.path for i in files])
    # for workbook in workbooks:
    sheet = None
    try:
        sheet = get_needed_sheet(get_workbook(path))
    except:
        print(f'Cant get sheet from {path}')
    if sheet is not None:
        return get_structure(PROJECT_STRUCTURES.ICJDATASTRUCT, \
            (path, sheet.name, get_data_from_icj(sheet)))
    else:
        pass

    # sheets = get_icj_sheets(workbooks)
    # for s in sheets:
        # data.append(namedtuple(*PROJECT_STRUCTURES.DS)._make((s.name, get_data_from_icj(s))))
    # return data


def getICJDatas(struct):
    out = []
    files = list(filter(lambda a: '.xls' in a.ext and 'ЖВК' in a.name and '~' not in a.name, sorted(struct, key=lambda a: a.date.c)))
    for file in files:
        out.append(getICJData(file.path))
    return out


def getData(text, icjdatastruct):
    out = []
    for item in icjdatastruct:
        if text in str(item.book):
                for _d in item.data:
                    out.append([_d.obj, _d.cert, _d.qa, _d.qp])
    return out


def getDate(text, icjdatastruct):
    out = []
    for item in icjdatastruct:
        if text in str(item.book):
                for _d in item.data:
                    out.append([_d.date])
    return max(out)
    

def injectICJData(data):
    book = xlwings.books.active
    name = book.name.split('.')[0]
    xlwings.books.active.sheets['Для ТСН'].select()
    xlwings.books.active.sheets['Для ТСН'].range('a5').select()
    xlwings.books.active.selection.value = getData(name, data)
    xlwings.books.active.sheets['DB'].select()
    xlwings.books.active.sheets['DB'].range('b3').select()
    xlwings.books.active.selection.value = getDate(name, data)


def getObjs(data):
    pattern = r'\d{5} *[eEеЕ]'
    regex = re.compile(pattern)
    out = {}
    for d in data:
        if d is not None:
            proj = str(d.book).split('\\')[-3].split('_')[0]
            for _d in d.data:
                for match in regex.finditer(_d.obj):
                    s = match.start()
                    e = match.end()
                    emid = _d.obj[s:e-1]
                    obj = _d.obj[0:s].strip()
                    cert = _d.cert
                    proj_emid = proj + '-' + emid + 'E'
                    # print(f'{proj_emid} -> {obj} -> {cert}')
                    if proj_emid in out.keys():
                        out[proj_emid][0].append(obj)
                        out[proj_emid][1].append(cert)
                    else:
                        out[proj_emid] = [[obj], [cert]]
    for d in out:
        out[d] = [list(set(out[d][0])), list(set(out[d][1]))]
    return out


def createDB(name):
    handle = 'proj-emid obj cert'
    path = r'E:\Проекты\archive'
    sheet_name = 'ICJ_DB_' + name
    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = sheet_name
    for i, c in enumerate(handle.split()):
        ws.cell(1, i+1).value = c
    _dict = getObjs(get(path))
    for row, d in enumerate(_dict):
        ws.cell(row+2, 1).value = d
        obj = '\n'.join(_dict[d][0]).strip()
        ws.cell(row+2, 2).value = obj
        cert = '\n'.join(_dict[d][1]).strip()
        ws.cell(row+2, 3).value = cert
    wb.save(path + '\\' + sheet_name + '.xlsx')
    wb.close()
    
