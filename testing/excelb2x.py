import pyxlsb
import openpyxl
from pathlib import Path
# from kip.utils.Structures import get_structure
# path = Path(r'E:\Проекты\current\ЖВК №2100-0012FN2780-112-0190-26.xlsb')


def get_data_from_xlsb(path):
    out = {}
    xlsb = pyxlsb.open_workbook(path)
    for sheet_num, sheet_name in enumerate(xlsb.sheets):
        sheet = xlsb.get_sheet(sheet_num + 1)
        out[sheet_name] = []
        for cells_list in sheet.rows():
            for cell in cells_list:
                out[sheet_name].append((cell.r, cell.c, cell.v))
    xlsb.close()
    return out


def write_data_to_xlsx(data, path:Path):
    xlsx = openpyxl.Workbook()
    name = str(path.absolute()).replace('xlsb', 'xlsx')
    # sheets = xlsx.sheetnames
    # xlsx.remove(sheets[0])
    for sheet in data.keys():
        xlsx.create_sheet(sheet)
        sh = xlsx[sheet]
        for d in data[sheet]:
            sh.cell(d[0]+1, d[1]+1, d[2])
    xlsx.save(name)
    xlsx.close()
    # print(f'path of new file {name}')
    # print(f'path of new file {Path(name)}')
    return Path(name)