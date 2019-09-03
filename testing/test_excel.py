import openpyxl


def get_eht_remarks(filename, cable_column='f', remark_column='l', check=True):
    "Сведение замечений по линиям грейки"
    wb = openpyxl.load_workbook(filename)
    ws = wb.sheetnames
    if check:
        check_columns(wb, cable_column, remark_column)
    out = []
    for sh in ws:
        _sh = wb.get_sheet_by_name(sh)
        max_row = _sh.max_row
        if max_row > 50000:
            print(f'Very large size {sh}: {max_row}')
        for line in range(2, max_row + 1):
            cabel_number = _sh[cable_column + str(line)].value
            remark = _sh[remark_column + str(line)].value
            if cabel_number is not None:
                out.append(f'{sh}\t{cabel_number.upper()}\t{remark}')
    wb.close()
    return out


def check_columns(workbook, cable_column='f', remark_column='l'):
    for sh in workbook.worksheets:
        print(f'Size of sheet {sh}: {sh.max_column}x{sh.max_row}')
        print(sh, end=' ')
        for i in sh.iter_cols():
            if i[0].column_letter.lower() in [i.lower() for i in (cable_column, remark_column)]:
                print(f'{i[0].column_letter} {i[0].value}', end=' ')
        print()
